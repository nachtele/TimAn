import os
from re import T
import numpy as np
import math
import matplotlib.pyplot as plt
from matplotlib import patches
import matplotlib.gridspec as gridspec
import openpyxl
from openpyxl.drawing.image import Image

def engFmt(num):
    if num == 0.0:
        return('0.0')
    e = (math.log(abs(num), 1000) // 1) * 3
    m = num / (10 ** e)
    fstr = f'{m:.3f}E{e:+.0f}' if e < -3 else f'{m:.6f}E{e:+.0f}'
    return(fstr)

ap = dict(color = 'gray', alpha = 0.5, width = 0, headwidth = 0, shrink = 0.0)

class makeFig:
    def __init__(self, tim, dat, tas, chs):
        self.aplist = []    #再設定時に削除するannotateとpatchリスト
        self.hllist = []    #再設定時に削除するhlineリスト
        nTr = len(dat)
        self.fig = plt.figure(figsize=tas.figsize, dpi=tas.dpi)
        gs0 = gridspec.GridSpec(len(tas.hratios), 1, height_ratios=tas.hratios)
        gs1 = [gs0[g].subgridspec(nTr, 1, hspace=0) for g in range(2)]
        self.ax = [[self.fig.add_subplot(gs1[g][tr]) for tr in range(nTr)] for g in range(2)]
        for i, d in enumerate(dat):
            for g in range(2):
                self.ax[g][i].set_title(chs[0][i], x=-0.06, y=0.2, loc='right', verticalalignment='center')
                self.ax[g][i].plot(tim, d, color=chs[1][i])
                self.ax[g][i].set_ylim(tas.ylim[g])
                self.ax[g][i].set_yticks(tas.yticks[g])
                if i > 0: self.ax[g][i].sharex(self.ax[g][i-1])
                if i < nTr-1: self.ax[g][i].tick_params(labelbottom=False)
        self.ax.append([self.fig.add_subplot(gs0[2])])
        self.ax[2][0].set_ylim(tas.ylim[2])
        self.ax[2][0].sharex(self.ax[1][0])
        self.ax[2][0].tick_params(labelbottom=False)
        self.ax[2][0].axis('off')
    def setStr(self, ta, strSettings):
        for name, tv, str, tr, vp in strSettings:
            for t, v in zip(ta.ht[name], ta.hv[name]):
                if v == tv:
                    self.ax[0][tr].annotate(str, xy=(t, vp), xytext=(t, vp), ha='right')
    def hlines(self, ch):
        for anp in self.hllist: anp.remove()
        self.hllist[:] = []
        for [c, th] in ch:
            self.hllist.append(self.ax[1][c].axhline(y=th, xmin=0, xmax=1, color='gray', alpha=0.5))
        pass
    def ann(self, ch, v, tg, tas, lblx):
        for anp in self.aplist: anp.remove()
        self.aplist[:] = []
        for [c, th], ly, t in zip(ch, tas.lny, tg):
            self.aplist.append(self.ax[1][c].annotate(f'{th:.2f}', xy=(lblx, th), xytext=(4, 0), va='center', textcoords='offset pixels'))
            self.aplist.append(self.ax[2][0].annotate(engFmt(t), xy=(t, ly[0]), xytext=(t, ly[1]), annotation_clip=False, arrowprops=ap))
            for ax in self.ax[1][c:]:
                self.aplist.append(ax.axvline(x=t, ymin=0, ymax=1, color='gray', alpha=0.5))
        self.aplist.append(self.ax[2][0].annotate(engFmt(v), xy=(tg[1], tas.vty), xytext=(tg[1], tas.vty)))
        patch = patches.FancyArrow(x=tg[1], y=tas.vly, dx=v, dy=0, width=0, head_width=0.1, head_length=0.05*v, length_includes_head=True)
        self.ax[2][0].add_patch(patch)
        self.aplist.append(patch)
    def zoom(self, xmin, xmax):
        self.ax[1][0].set_xlim(xmin, xmax)
        for ax in self.ax[0]:
            for p in list(ax.patches): p.remove()
            ylim = ax.get_ylim()
            patch = patches.Rectangle(xy=(xmin,ylim[0]), width=xmax-xmin, height=ylim[1]-ylim[0], ec='steelblue', fc='steelblue')
            ax.add_patch(patch)

def zoomScales(llimit):
    ex = int(math.log10(llimit)//1)
    zs = np.array([float(str(a)+'e'+str(b)) for b in range(ex,-1) for a in [1,2,5]])
    return zs[np.searchsorted(zs, llimit):]

def plotTa(tim, dat, ta, tas, chs, zs, dstDir):
    plt.rcParams['font.family'] = 'MS Gothic'
    dispList = []
    afig= makeFig(tim, dat, tas, chs)
    afig.setStr(ta, tas.strs)
    os.mkdir(dstDir)
    dn = len(tas.disp)
    for di, [name, caption, ch, sel] in enumerate(tas.disp):
        plt.suptitle(caption)
        hv = np.array(ta.hv[name])
        sellist = []
        if len(hv) >0:   #要素があれば処理
            imax = hv.argmax()
            imin = hv.argmin()
            absmin = np.amin(np.abs(hv))
            zmin = zs[np.searchsorted(zs, absmin*6)]
            for s in sel:
                if s == 'all':
                    sellist = [[i, dstDir + '\\' + name + '_' + str(i) + '.png'] for i in range(len(hv))]
                    break
                elif s == 'min':
                    if len(sellist) == 0 or sellist[-1][0] != imin:
                        sellist.append([imin, dstDir + '\\' + name + '_' + str(imin) + '.png'])
                elif s == 'max':
                    if len(sellist) == 0 or sellist[-1][0] != imax:
                        sellist.append([imax, dstDir + '\\' + name + '_' + str(imax) + '.png'])
            afig.hlines(ch)
        dispList.append(sellist)
        sn = len(sellist)
        for si, [i, dstFile] in enumerate(sellist):
            v = hv[i]
            tg = ta.hg[name][i]
            z = zs[np.searchsorted(zs, abs(v)*2)] if abs(v)*2 > zmin else zmin
            xmax = (tg[0] + tg[1] + z) / 2
            xmin = xmax - z
            afig.zoom(xmin, xmax)
            afig.ann(ch, v, tg, tas, xmax)
            if tas.mainZoom:
                for mzmin, mzmax in tas.mainZoom:
                    if mzmin <= tg[0] and tg[1] <= mzmax:
                        afig.ax[0][0].set_xlim(xmin=mzmin)
                        afig.ax[0][0].set_xlim(xmax=mzmax)
                        break
            plt.savefig(dstFile)
            print('\r', di+1, '/', dn, ' ', name, ' ', si+1, '/', sn, end='')
    return dispList

def taXlsx(ta, tas, dispList, info, dstDir):
    wb = openpyxl.Workbook()
    #測定値一覧シート
    ws = wb.active
    ws.title = 'data'
    for i, [t, v, f] in enumerate(info):
        ws.cell(i+1, 1).value = t
        ws.cell(i+1, 2).value = v
        ws.cell(i+1, 2).number_format = f
    InitCell = [len(info)+3, 2]
    colInc = 4
    col = InitCell[1]
    for ds in tas.disp:
        row = InitCell[0]
        name = ds[0]
        cc = [name, ds[1], len(ta.hv[name])]
        for c in cc:
            ws.cell(row, col).value = c
            row += 1
        for ri, [v, g] in enumerate(zip(ta.hv[name], ta.hg[name])):
            ws.cell(row, col).value = ri
            cc = [v, g[0], g[1]]
            for ci, c in enumerate(cc):
                ws.cell(row, col+ci+1).value = c
                ws.cell(row, col+ci+1).number_format = '##0.000E+0'
                ws.column_dimensions[openpyxl.utils.get_column_letter(col+ci+1)].width = 10
            row += 1
        col += colInc

    #波形シート
    ws = wb.create_sheet('waveform')
    iScale = 0.50
    InitCell = [3,2]
    rowInc = 14
    colInc = 3
    row = InitCell[0]
    for ds, dl in zip(tas.disp, dispList):
        col = InitCell[1]
        name = ds[0]
        n = len(ta.hv[name])
        ws.cell(row, col).value = ds[1]
        if n == 0: continue
        for d in dl:
            i = d[0]
            cc = [
                ['i', i, '']
                ,['n', n, '']
                ,[name, ta.hv[name][i], '##0.000E+0']
                ,['t0', ta.hg[name][i][0], '##0.000E+0']
                ,['t1', ta.hg[name][i][1], '##0.000E+0']
            ]
            for ri, [t, v, f] in enumerate(cc):
                ws.cell(row+ri+1, col).value = t
                ws.cell(row+ri+1, col+1).value = v
                ws.cell(row+ri+1, col+1).number_format = f
            for i,w in enumerate([14, 11, 44]):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col+i)].width = w
            img= Image(d[1])
            img.height *= iScale
            img.width *= iScale
            cell = ws.cell(row+1, col+2).coordinate
            ws.add_image(img, cell)
            col += colInc
        row += rowInc
    wb.save(dstDir + '.xlsx')
